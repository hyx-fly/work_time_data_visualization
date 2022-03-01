import os
import time
import logging
import webbrowser
from openpyxl import load_workbook
from datetime import datetime
from win32com.client import Dispatch


def print_execute_time(func):
    """
    获取程序执行时间的一个装饰器

    """

    def wrapper(*args, **kwargs):
        start_time = time.time()
        func_return = func(*args, **kwargs)
        end_time = time.time()
        print_time = round((end_time - start_time), 2)
        print(f'完成用时{print_time}秒')
        return func_return

    return wrapper


def hyx_log(text, log_name='工作日志'):
    """
    生成日志文件

    :param text: 日志文本
    :param log_name: 日志文件名
    """
    log_format = "%(asctime)s - %(levelname)s - %(message)s"
    date_format = "%Y/%m/%d %H:%M:%S %p"
    logging.basicConfig(filename=f'./data/{log_name}.log', level=logging.INFO, format=log_format, datefmt=date_format)
    print(text)
    logging.info(text)


def time_out(off_time: str):
    """
    程序过期时间

    :param off_time: 格式:1990/01/01
    """
    now_time = datetime.now().strftime('%Y/%m/%d')
    out_of_num = (datetime.strptime(off_time, "%Y/%m/%d") - datetime.strptime(now_time, "%Y/%m/%d")).days
    # print(out_of_num)
    if out_of_num < 0:
        exit()


def the_password(password: str):
    """
    设置一个启动密码

    :param password: 设置的密码
    """
    while True:
        user_password = input('请输入启动密码：')
        if user_password == password:
            break
        else:
            print('密码错误！')


def gen_file_list(path) -> list:
    """
    遍历路径，生成一个路径下所有文件的路径列表。

    :param path: 路径地址
    :return: 路径下所有文件的列表
    """
    result = []
    for file_name in os.listdir(path):
        file_path = ''.join([path, file_name])
        result.append(file_path)

    # print(result)
    return result


def copy_new_xlsx(path: str, save_path: str):
    """
    拷贝模板xlsx文件到指定地址。

    :param path: 模板文件地址
    :param save_path: 模板文件地址
    """
    wb = load_workbook(path)
    wb.save(save_path)


def position_top_2title(sheet, title_1='工号', title_2='预测劳动时间', max_row=3):
    """
    从1-max_row行中查找指定标题的row值。

    :param sheet: 一个工作表对象
    :param title_1: 标题1
    :param title_2: 标题2
    :param max_row: 查找的最大row值
    :return: 标题1/2所在row的元组
    """
    result = []
    for row in range(1, max_row + 1):
        for col in sheet[row]:
            if col.value == title_1:
                result.append(row)
                result.append(col.column)
            elif col.value == title_2:
                result.append(col.column)

    return tuple(result)


def position_any_title(sheet, title: str):
    """
    查找工作表中任意单元格（value）的所在row，col，定位失败会返回（0，0）

    :param sheet: 一个工作表对象
    :param title: 任意的单元格value
    :return: 单元格位置（row，col）元祖
    """
    result = (0, 0)
    for row in range(1, sheet.max_row + 1):
        for col in sheet[row]:
            if col.value == title:
                result = (row, col.column)
                return result

    print(f'定位失败,未找到{title}')
    return result


def gen_value_list(path):
    """
    从指定xlsx文件中查找指定2个标题栏下所有的值，并组成对应元组。

    :param path: xlsx文件地址
    :return: 多个（标题1.value，标题2.value）组成的元祖
    """
    result = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb:
        location = position_top_2title(sheet)
        row_start, id_col, work_time_col = location
        for row in range(row_start + 1, sheet.max_row + 1):
            value1 = sheet.cell(row, id_col).value
            value2 = sheet.cell(row, work_time_col).value
            if bool(value1) and bool(value2):
                # if '#' in str(value1):
                #     hyx_log(f'{path}>>[{sheet.title}]>>[第{row}行{id_col}列]单元格>>信息有误')

                if '/' in str(value1):  # 双班
                    for v in value1.split(sep='/'):
                        result.append((int(v), value2))

                else:  # 单班
                    result.append((int(value1), value2))

    print(path, '=>合计找到', len(result), '人')
    return tuple(result)


def save_work_time_data(path, u_dict: dict, today: str):
    """
    保存特定数据到模板xlsx文件中。

    :param path: 文件地址
    :param u_dict: 数据字典
    :param today: 指定日期（某日的日期，例子：1日）
    """
    wb = load_workbook(path)
    sheet = wb['统计']
    col = sheet['B']
    for cell in col:
        if cell.value == '工号' or cell.value is None:
            continue
        try:
            value = u_dict[cell.value]
            u_row = cell.row
            u_col = position_any_title(sheet, today)[1]
            # print(value, u_row, u_col)
            sheet.cell(u_row, u_col).value = value
        except KeyError:
            continue

    wb.save(path)


def read_total_work_time(path, sheet_name, col_str) -> list:
    """
    从xlsx数据文件中读取指定数据。

    :param path: 文件地址
    :param sheet_name: 数据表名称
    :param col_str: 指定列的英文代码
    :return: 数据列表
    """
    result = []
    wb = load_workbook(path, data_only=True)
    sheet = wb[sheet_name]
    col = sheet[col_str]
    for cell in col:
        if type(cell.value) is not str:
            name = sheet.cell(cell.row, 1).value
            u_id = sheet.cell(cell.row, 2).value
            person = f'{name}{u_id}'
            total_work_time_data = cell.value
            if total_work_time_data is not None:
                total_work_time = round(total_work_time_data.total_seconds() / 3600)
                result.append((person, total_work_time))

    return result


def read_total_work_time_one_day(path, stop_num, title_str: str):
    """
    从xlsx数据文件中读取某行（row）劳时数据。

    :param path: xlsx地址
    :param stop_num: 停止读取的次数
    :param title_str: 指定行的标题value
    :return: 数据元祖
    """
    wb = load_workbook(path, data_only=True)
    sheet = wb['统计']
    row = position_any_title(sheet, title_str)[0]

    result = []
    if bool(row):
        sum_time_list = []
        for cell in sheet[row]:
            if stop_num - len(sum_time_list) == 1 or cell.value == 0:
                break
            elif cell.value is not None and type(cell.value) is not str:
                sun_time_data = cell.value
                sum_time = round(sun_time_data.total_seconds() / 3600)
                sum_time_list.append(sum_time)

        sum_time_index = [index_date for index_date in range(1, len(sum_time_list) + 1)]
        result = list(zip(sum_time_index, sum_time_list))

    return tuple(result)


def open_xlsx(path, is_show=False):
    """
    使用wps或者office打开xlsx文件。
    :param path: xlsx文件地址
    :param is_show: 是否显示，默认不显示
    """
    path = os.path.abspath(path)
    try:
        # 用wps打开
        xl_app = Dispatch('Ket.Application')
        xl_app.Visible = is_show
        xl_book = xl_app.Workbooks.Open(path)
        xl_book.Save()
        if is_show is False:
            xl_book.Close()
            xl_app.Quit()

    except BaseException:
        # 用office打开
        xl_app = Dispatch('Excel.Application')
        xl_app.Visible = is_show
        xl_book = xl_app.Workbooks.Open(path)
        xl_book.Save()
        if is_show is False:
            xl_book.Close()
            xl_app.Quit()


def work_time_type_count(data, count_type=('>零<劳时', '劳时过低', '劳时正常', '劳时过高'),
                         up_data=180, low_data=60):
    """
    计算达到阈值的数据数量。

    :param data: 数据组
    :param count_type: 分类
    :param up_data: 上限阈值
    :param low_data: 下限阈值
    :return: 计数元祖
    """
    time_type = count_type
    zero = 0
    low = 0
    medium = 0
    high = 0
    for data in data:
        if data[1] > up_data:
            high += 1
        elif 0 < data[1] < low_data:
            low += 1
        elif int(data[1]) == 0:
            zero += 1
        else:
            medium += 1

    result = list(zip(time_type, (zero, low, medium, high)))

    return tuple(result)


def open_html(url_path: str):
    """
    调用默认浏览器打开网址。

    :param url_path: 网页地址
    """
    url_path = os.path.abspath(url_path)
    webbrowser.open(os.path.abspath(url_path))


def gen_work_time_main():
    """
    数据格式化主函数。
    :return 数据文件地址，当前月份int，当前日期int
    """

    # 当前日期
    localtime = time.localtime(time.time())  # 系统时间
    the_month_num = localtime.tm_mon  # 月
    today_num = localtime.tm_mday  # 日

    # 文件及数据路径
    data_path = f'./data/'
    mon_path = f'{data_path}{the_month_num}月/'
    save_path = f'./data/output/劳时干预数据_{the_month_num}月.xlsx'

    # 创建预警数据文件
    if os.path.exists(save_path) is False:
        copy_new_xlsx(f'{data_path}劳时干预数据_模板.xlsx', save_path)
        for index in range(1, today_num):
            # 当日计划文件
            today_file_path = mon_path + f'出勤计划{index}日.xlsx'

            try:
                # 存储数据
                data_dict = dict(gen_value_list(today_file_path))
                save_work_time_data(save_path, data_dict, f'{index}日')
            except BaseException as msg:
                hyx_log(f'错误信息:{msg}')

    else:
        # 当日计划文件
        today_file_path = f'{mon_path}出勤计划{today_num - 1}日.xlsx'

        try:
            # 存储数据
            data_dict = dict(gen_value_list(today_file_path))
            save_work_time_data(save_path, data_dict, f'{today_num - 1}+ 日')
        except BaseException as msg:
            hyx_log(f'错误信息:{msg}')

    return save_path, the_month_num, today_num
